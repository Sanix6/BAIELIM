from django.shortcuts import render
from rest_framework import viewsets, permissions, status, generics
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse

from category.models import File
from user import models, serializers, filters
from user.functions import create_store_inOneC, create_stores_inOneC
from rest_framework import pagination
import openpyxl
from django.core.files.base import File as DjangoFile
from tempfile import NamedTemporaryFile
from user.models import Store


class AdministratorViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    queryset = models.ModelAdmin.objects.all()
    serializer_class = serializers.AdminSerializer


class ManagerViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    queryset = models.Manager.objects.all()
    serializer_class = serializers.ManagerSerializer


class AgentViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    queryset = models.Agent.objects.all().order_by('name')
    serializer_class = serializers.AgentSerializer
    filter_backends = (DjangoFilterBackend, OrderingFilter, SearchFilter)
    filter_class = filters.AgentFilter
    search_fields = ('name', 'phoneNumber')

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return serializers.AgentSerializerGet
        else:
            return serializers.AgentSerializer


class StoreViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    queryset = models.Store.objects.all().order_by('name')
    serializer_class = serializers.StoreSerializer
    filter_backends = (DjangoFilterBackend, OrderingFilter, SearchFilter)
    filter_class = filters.StoreFilter
    search_fields = ('name', 'contactName', 'phoneNumber', 'phoneNumber1', 'phoneNumber2', 'phoneNumber3', 'address')

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return serializers.StoreSerializerGet
        else:
            return serializers.StoreSerializer

    def create(self, request, *args, **kwargs):

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        saved_data = serializer.save()

        if not request.data.get('oneC_code'):
            saved_data.oneC_code = request.data.get('login')
            saved_data.save()

        response_data = serializers.StoreSerializerGet(saved_data).data
        # name = saved_data.name
        # user_id = saved_data.id
        # phoneNumber = saved_data.phoneNumber
        # address = saved_data.address
        # agent = models.Agent.objects.filter(pk=saved_data.store_agent).first()
        # if agent:
        #     agent_name = agent.name
        # else:
        #     agent_name = ''
        headers = self.get_success_headers(serializer.data)
        # oneC = create_store_inOneC(user_id, name, agent_name, phoneNumber, address)
        # response_data = {
        #     'oneC': oneC,
        #     'object': serializer.data,
        # }
        # if oneC['code'][0] == 200:
        #     try:
        #         store = models.Store.objects.get(pk=user_id)
        #         code = oneC['data']['Покупатели'][0]['Код']
        #         store.oneC_code = code
        #         store.save()
        #     except models.Store.DoesNotExist:
        #         raise {'error': 'error'}
        return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        self.perform_update(serializer)
        saved_data = self.get_object()
        name = saved_data.name
        user_id = saved_data.id
        phoneNumber = saved_data.phoneNumber
        address = saved_data.address
        agent = models.Agent.objects.filter(pk=saved_data.store_agent).first()
        if agent:
            agent_name = agent.name
        else:
            agent_name = ''
        headers = self.get_success_headers(serializer.data)
        oneC = create_store_inOneC(user_id, name, agent_name, phoneNumber, address)
        response_data = {
            'oneC': oneC,
            'object': serializer.data,
        }
        if oneC['code'][0] == 200:
            try:
                store = models.Store.objects.get(pk=user_id)
                code = oneC['data']['Покупатели'][0]['Код']
                store.oneC_code = code
                store.save()
            except models.Store.DoesNotExist:
                raise {'error': 'error'}


        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)


class DriverViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    queryset = models.Driver.objects.all()
    serializer_class = serializers.DriverSerializer
    filter_backends = (DjangoFilterBackend, OrderingFilter, SearchFilter)
    # filter_class = filters.DriverFilter
    search_fields = ('name', )


class DayPlanViewSet(viewsets.ModelViewSet):
    queryset = models.DayPlan.objects.all()
    serializer_class = serializers.DayPlanSerializer
    filter_backends = (DjangoFilterBackend, OrderingFilter, SearchFilter)
    filter_class = filters.DayPlanFilter
    search_fields = ('stores__name', 'stores__phoneNumber', 'stores__address', 'agent__name', 'agent__phoneNumber')

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return serializers.DayPlanSerializerGet
        else:
            return serializers.DayPlanSerializer


class StorePlanViewSet(viewsets.ModelViewSet):
    queryset = models.PlanStore.objects.all()
    serializer_class = serializers.PlanStoreSerializer
    filter_backends = (DjangoFilterBackend, OrderingFilter, SearchFilter)
    filter_class = filters.StorePlanFilter

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return serializers.PlanStoreSerializerGet
        else:
            return serializers.PlanStoreSerializer

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        response_serializer = serializers.PlanStoreSerializerGet(instance)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(response_serializer.data)


class CustomPagination(pagination.PageNumberPagination):
    page_size = 300
    page_size_query_param = 'page_size'
    max_page_size = 100
    page_query_param = 'page'


class GetStoresView(APIView):
    permission_classes = [permissions.AllowAny]
    pagination_class = CustomPagination

    def get(self, request):
        queryset = models.Store.objects.filter(lat__gt=0, lon__gt=0)
        paginator = self.pagination_class()
        serializer = serializers.StoreSerializerForGet(queryset, many=True)
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = serializers.StoreSerializerForGet(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        serializer = serializers.StoreSerializerForGet(queryset, many=True)
        return Response(serializer.data)


class NewDayPlanView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        day_plans = models.DayPlan.objects.filter(status='new')
        plans_id = []
        for day_plan in day_plans:
            new_day_plan = models.DayPlan.objects.create(
                day=day_plan.day, agent=day_plan.agent, status='new', driver=day_plan.driver,
                dateVisit=day_plan.dateVisit
            )
            new_day_plan.stores_plan.set(day_plan.stores_plan.all())
            new_day_plan.save()
            day_plan.status = 'archive'
            day_plan.save()
            plans_id.append(day_plan.id)
        return Response({'status': 'ok', 'day_plans': plans_id})


class GetStoreUntilJuneView(APIView):
    permission_classes = [permissions.AllowAny]
    pagination_class = CustomPagination

    def get(self, request):
        queryset = models.Store.objects.all().filter(dateCreated__date__lte='2024-06-05')
        queryset = queryset.order_by('-dateCreated')
        paginator = self.pagination_class()
        serializer = serializers.StoreSerializerJune(queryset, many=True)
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = serializers.StoreSerializerJune(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        serializer = serializers.StoreSerializerJune(queryset, many=True)
        return Response(serializer.data)

# class GetJulyStoresView(APIView):
#     permission_classes = [permissions.AllowAny]
#     pagination_class = CustomPaginationStore()
#
#     def list(self, request, *args, **kwargs):
#         queryset = models.Store.objects.all()
#
#         page = self.paginate_queryset(queryset)
#         if page is not None:
#             serializer = serializers.StoreSerializerGet(page, many=True)
#             return self.get_paginated_response(serializer.data)
#
#         serializer = serializers.StoreSerializerGet(queryset, many=True)
#         return Response(serializer.data)


class GetStoreForOneCView(APIView):
    permission_classes = [permissions.AllowAny]
    pagination_class = CustomPagination

    def get(self, request):
        queryset = models.Store.objects.all().filter(dateCreated__date__gte='2024-09-04')
        paginator = self.pagination_class()
        serializer = serializers.OneCStoreSerializer(queryset, many=True)
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = serializers.OneCStoreSerializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        serializer = serializers.OneCStoreSerializer(queryset, many=True)
        return Response(serializer.data)


class ChangePasswordWithoutOldPasswordView(generics.UpdateAPIView):
    """
    An endpoint for changing password.
    """
    serializer_class = serializers.ChangePasswordWithoutOldPassowrdSerializer
    model = models.User
    permission_classes = (permissions.AllowAny,)

    def get_object(self, queryset=None):
        obj = self.request.user
        return obj

    def update(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)

        # self.object = self.get_object()

        if serializer.is_valid():
            user_id = models.User.objects.get(pk=serializer.data.get('user_id'))
            self.object = user_id
            # set_password also hashes the password that the user will get
            self.object.set_password(serializer.data.get("new_password"))
            self.object.save()
            response = {
                'status': 'success',
                'code': status.HTTP_200_OK,
                'message': 'Password updated successfully',
                'data': []
            }

            return Response(response)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.pagination import PageNumberPagination


class StorePagination(PageNumberPagination):
    page_size = 300


class SynchronizeStoresView(APIView):
    permission_classes = (permissions.AllowAny, )
    pagination_class = StorePagination

    def post(self,request):
        stores = models.Store.objects.filter(dateCreated__lte='2024-08-01')

        paginator = self.pagination_class()
        paginated_stores = paginator.paginate_queryset(stores, request)

        send_stores = []
        for store in paginated_stores:
            guid = '00000000-0000-0000-0000-'
            txt = str(store.id).zfill(12)
            guid = guid + txt
            int_code = 10000 + store.id
            code = store.name[0:3] + str(int_code)
            if len(code) < 9:
                result_string = code.ljust(9, 'x')
            else:
                result_string = code[:9]
            send_store = {
                "Guid": guid,
                "Код": result_string,
                "Наименование": f"{store.name}, {store.address}, {store.phoneNumber}",
                "Комментарий": f'{store.store_agent}'
            }
            send_stores.append(send_store)

        data = {
            "Table": "Покупатели",
            "Покупатели": send_stores
        }

        return paginator.get_paginated_response(data)


class EptaView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, format=None):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Stores"

        # Устанавливаем заголовки столбцов
        headers = ["ID", "Название", "Адрес", "Агент (fullname)"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Получаем данные из модели Store
        stores = models.Store.objects.filter(order__isnull=False).distinct()
        # stores = stores.select_related('store_agent').all()

        # Добавляем строки с данными
        for row_num, store in enumerate(stores, 2):
            sheet.cell(row=row_num, column=1, value=store.id)
            sheet.cell(row=row_num, column=2, value=store.name)
            sheet.cell(row=row_num, column=3, value=store.address)
            sheet.cell(row=row_num, column=4, value=store.store_agent.name if store.store_agent else "Нет агента")

        # Создаем временный файл для сохранения
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)  # Возвращаем курсор в начало файла
            django_file = DjangoFile(tmp, name="stores.xlsx")

            # Сохраняем файл в модель File
            file_instance = File.objects.create(
                title="Экспорт магазинов",
                file=django_file
            )

        return Response({"data": file_instance.id})


class UpdateStoreOneCCodeView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'error': 'Excel file is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Открытие Excel-файла
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Чтение строк из Excel-файла
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
                name, onecode, agent = row

                if not name or not onecode or not agent:
                    continue  # Пропускаем строки с пустыми значениями

                # Фильтрация по имени магазина
                stores = models.Store.objects.filter(name=name)
                if stores.count() == 1:
                    # Если найден один магазин, обновляем oneC_code
                    store = stores.first()
                    store.oneC_code = onecode
                    store.save()
                elif stores.count() > 1:
                    # Если магазинов больше одного, фильтруем по store_agent
                    agent_instance = models.Agent.objects.filter(name=agent).first()
                    if agent_instance:
                        store = stores.filter(store_agent=agent_instance).first()
                        if store:
                            store.oneC_code = onecode
                            store.save()

            return Response({'message': 'Stores updated successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from openpyxl.styles import Font

def export_to_excel(request):
    # Создаем книгу и активный лист
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Stores Data'

    # Задаем заголовки
    headers = ['GUID', 'Agent', 'Region', '1C Code']
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Получаем данные из базы
    stores = Store.objects.select_related('store_agent', 'region')

    # Добавляем данные в таблицу
    for row_num, store in enumerate(stores, start=2):
        sheet.cell(row=row_num, column=1).value = store.guid
        sheet.cell(row=row_num, column=2).value = str(store.store_agent) if store.store_agent else 'N/A'
        sheet.cell(row=row_num, column=3).value = str(store.region) if store.region else 'N/A'
        sheet.cell(row=row_num, column=4).value = store.oneC_code if hasattr(store, 'oneC_code') else 'N/A'

    # Формируем HTTP-ответ для загрузки файла
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stores_data.xlsx"'
    workbook.save(response)

    return response